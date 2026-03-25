import pandas as pd
import json
import re
import os
import openpyxl  # (used by pandas for xlsx)
import requests
import pygsheets
import glob
import traceback
import logging
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parent
VAR = ROOT / "var"
VAR.mkdir(exist_ok=True)

CACHE_FILE = VAR / "cache.json"
CREDENTIALS = ROOT / "auth.json"
BASE_NAME = "analytics"

SHEET_DATE_CACHE: dict = {}
DATE_PATTERN = re.compile(r"\((\d{2}\.\d{2}\.\d{4})\)")

for ext in (".xls", ".xlsx", ".xlsx.xls", ".xls.xlsx"):
    fp = VAR / f"{BASE_NAME}{ext}"
    if fp.exists():
        fp.unlink()

SERVICE_ACCOUNT_FILE = str(CREDENTIALS)
SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

try:
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    drive = build("drive", "v3", credentials=creds)
    logger.info(f"Google Drive API инициализирована")
except Exception as e:
    logger.error(f"Ошибка инициализации Google Drive API: {e}")
    logger.error(traceback.format_exc())
    raise

try:
    gc = pygsheets.authorize(service_file=SERVICE_ACCOUNT_FILE)
    spreadsheet_id = gc.open(BASE_NAME).id
    logger.info(f"Google Sheets инициализирована, spreadsheet_id={spreadsheet_id}")
except Exception as e:
    logger.error(f"Ошибка инициализации Google Sheets: {e}")
    logger.error(traceback.format_exc())
    raise

excel_path = VAR / f"{BASE_NAME}.xlsx"
try:
    logger.info(f"Экспортируем XLSX из Google Sheets...")
    request = drive.files().export_media(
        fileId=spreadsheet_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    with open(excel_path, "wb") as f:
        f.write(request.execute())
    logger.info(f"Файл аналитики успешно выгружен: {excel_path}")
except Exception as e:
    logger.error(f"Ошибка при экспорте XLSX: {e}")
    logger.error(traceback.format_exc())
    raise

def load_cache():
    if CACHE_FILE.exists():
        with CACHE_FILE.open("r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_cache(cache: dict):
    with CACHE_FILE.open("w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

def parse_old(material: dict, sheets: dict) -> dict | None:
    if material.get("parserType") != "old":
        return None
    
    sheet_name = material["sheet"]
    try:
        df = sheets[sheet_name]
    except KeyError:
        logger.error(f"parse_old: Лист '{sheet_name}' не найден в sheets. Доступные: {list(sheets.keys())}")
        return None
    except Exception as e:
        logger.error(f"parse_old: Ошибка при доступе к листу '{sheet_name}': {e}")
        return None

    try:
        date_col = material["date"]["header"]
        material_id = material["materialId"]

        props = material["properties"]
        prop_headers = [p["header"] for p in props]

        cols_to_check = [(material["materialName"], h) for h in prop_headers]

        df_filtered = df.dropna(subset=cols_to_check, how="all")
        if df_filtered.empty:
            logger.warning(f"parse_old: Все строки пусты для {material_id}")
            return None

        last_row = df_filtered.iloc[-1]
        
        if isinstance(date_col, tuple):
            last_date = last_row[date_col]
        else:
            try:
                last_date = last_row[date_col]
            except KeyError:
                matching_cols = [col for col in df.columns if col[0] == date_col]
                if matching_cols:
                    last_date = last_row[matching_cols[0]]
                else:
                    logger.error(f"parse_old: Колонка '{date_col}' не найдена для {material_id}")
                    return None
        
        if isinstance(last_date, pd.Series):
            last_date = last_date.iloc[0]
        if hasattr(last_date, "strftime"):
            last_date = last_date.strftime(material["date"]["format"])
        else:
            last_date = str(last_date)

        property_values = []
        for p in props:
            col = (material["materialName"], p["header"])
            if col in df.columns:
                val = last_row[col]
                if pd.notna(val):
                    property_values.append({
                        "propertyId": p["propertyId"],
                        "value": str(val)
                    })

        return {
            "materialId": material_id,
            "dateValues": [{"date": last_date, "propertyValues": property_values}]
        }
    except Exception as e:
        logger.error(f"parse_old: Ошибка обработки материала {material.get('materialId')}: {e}")
        logger.error(traceback.format_exc())
        return None

def parse_new(material: dict, sheets: dict) -> dict | None:
    if material.get("parserType") != "new":
        return None

    sheet_name = material["sheet"]
    try:
        df = sheets[sheet_name]
    except KeyError:
        logger.error(f"parse_new: Лист '{sheet_name}' не найден в sheets. Доступные: {list(sheets.keys())}")
        return None
    except Exception as e:
        logger.error(f"parse_new: Ошибка при доступе к листу '{sheet_name}': {e}")
        logger.error(traceback.format_exc())
        return None

    try:
        df_simple = df.copy(deep=False)
        df_simple.columns = df_simple.columns.get_level_values(0)

        last_date = None
        last_date_col = None
        cached = SHEET_DATE_CACHE.get(sheet_name)
        if cached is not None:
            raw_date_str, cached_col = cached
            last_date = datetime.strptime(raw_date_str, "%d.%m.%Y").strftime(material["date"]["format"])
            last_date_col = cached_col
        else:
            for col in reversed(df.columns):
                if isinstance(col[0], str):
                    match = DATE_PATTERN.search(col[0])
                    if match and col[1] == "Цена":
                        if df[col].notna().any():
                            raw = match.group(1)
                            last_date = datetime.strptime(raw, "%d.%m.%Y").strftime(material["date"]["format"])
                            last_date_col = col[0]
                            SHEET_DATE_CACHE[sheet_name] = (raw, last_date_col)
                            break

        if not last_date:
            logger.warning(f"parse_new: Не найдена дата для {material['materialId']}")
            return None

        f = material["rowFilter"]
        row_idx = df_simple[
            (df_simple["Страна"].astype(str) == f["country"]) &
            (df_simple["Порт"].astype(str) == f["port"]) &
            (df_simple["Вид материала"].astype(str) == f["type"])
        ].index

        if row_idx.empty:
            logger.warning(f"parse_new: Не найдено: {f}")
            return None

        row_idx = row_idx[0]

        props = []
        for p in material["properties"]:
            col = (last_date_col, p["column"])
            value = df.loc[row_idx, col]
            props.append({
                "propertyId": p["propertyId"],
                "value": str(value) if pd.notna(value) else None
            })

        return {
            "materialId": material["materialId"],
            "dateValues": [{"date": last_date, "propertyValues": props}]
        }
    except Exception as e:
        logger.error(f"parse_new: Ошибка обработки материала {material.get('materialId')}: {e}")
        logger.error(traceback.format_exc())
        return None


def parse_excel_formulas(material: dict, excel_path: str, wb_formulas=None, wb_values=None) -> dict | None:
    parser_type = material.get("parserType")
    if parser_type not in ("formula", "value"):
        logger.warning(f"parse_excel_formulas: Неизвестный parserType '{parser_type}' для '{material.get('materialName')}'")
        return None

    properties = material.get("properties", [])
    if not properties:
        logger.warning(f"parse_excel_formulas: Нет свойств для '{material.get('materialName')}'")
        return None

    close_books = False
    if wb_formulas is None or wb_values is None:
        wb_formulas = openpyxl.load_workbook(excel_path, data_only=False)
        wb_values = openpyxl.load_workbook(excel_path, data_only=True)
        close_books = True

    sheet_name = material.get("sheet")

    if sheet_name not in wb_values.sheetnames:
        logger.error(f"parse_excel_formulas: Лист '{sheet_name}' не найден. Доступные: {wb_values.sheetnames}")
        if close_books:
            wb_formulas.close()
            wb_values.close()
        return None

    ws_v = wb_values[sheet_name]
    ws_f = wb_formulas[sheet_name]

    try:
        ref_col = next((p["colLetter"] for p in properties if p.get("name") == "med"), properties[0]["colLetter"])
        last_row = None
        try:
            col_idx = column_index_from_string(ref_col)
            col_vals_iter = ws_v.iter_cols(min_col=col_idx, max_col=col_idx, values_only=True)
            col_vals = next(col_vals_iter)
            for i, val in enumerate(reversed(col_vals), start=0):
                if val not in (None, ""):
                    last_row = ws_v.max_row - i
                    break
        except Exception as e:
            for row in range(ws_v.max_row, 0, -1):
                cell = ws_v[f"{ref_col}{row}"]
                if cell.value not in (None, ""):
                    last_row = row
                    break

        if last_row is None:
            logger.warning(f"parse_excel_formulas: Нет данных для '{material.get('materialName')}'")
            if close_books:
                wb_formulas.close()
                wb_values.close()
            return None

        date_val = ws_v[f"A{last_row}"].value
        property_values = []

        if parser_type == "formula":
            med_prop = next((p for p in properties if p.get("name") == "med"), None)
            if not med_prop:
                logger.error(f"parse_excel_formulas: Нет поля med для '{material.get('materialName')}'")
                if close_books:
                    wb_formulas.close()
                    wb_values.close()
                return None

            cell_addr = f"{med_prop['colLetter']}{last_row}"
            cell_formula = ws_f[cell_addr].value
            cell_val = ws_v[cell_addr].value

            min_val = max_val = None
            if isinstance(cell_formula, str) and cell_formula.startswith("="):
                numbers = re.findall(r"\d+(?:\.\d+)?", cell_formula)
                if len(numbers) >= 2:
                    min_val, max_val = float(numbers[0]), float(numbers[1])

            if isinstance(cell_val, (int, float)):
                cell_val = float(cell_val)

            for prop in properties:
                if prop["name"] == "min":
                    val = min_val
                elif prop["name"] == "max":
                    val = max_val
                elif prop["name"] == "med":
                    val = cell_val
                else:
                    val = None
                
                value_str = str(val) if val is not None else None

                property_values.append({
                    "propertyId": prop["propertyId"],
                    "value": value_str
                })

        else:  # parserType == "value"
            for prop in properties:
                col = prop["colLetter"]
                raw_val = ws_v[f"{col}{last_row}"].value
                if isinstance(raw_val, (int, float)):
                    value_str = str(float(raw_val))
                elif raw_val is None:
                    value_str = None
                elif hasattr(raw_val, "strftime"):
                    value_str = raw_val.strftime("%Y-%m-%d")
                else:
                    value_str = str(raw_val)
                property_values.append({
                    "propertyId": prop["propertyId"],
                    "value": value_str
                })

        return {
            "materialId": material["materialId"],
            "dateValues": [
                {
                    "date": date_val.strftime("%d.%m.%Y") if hasattr(date_val, "strftime") else str(date_val) if date_val else None,
                    "propertyValues": property_values
                }
            ]
        }
    except Exception as e:
        logger.error(f"parse_excel_formulas: Ошибка обработки материала {material.get('materialId')}: {e}")
        logger.error(traceback.format_exc())
        if close_books:
            wb_formulas.close()
            wb_values.close()
        return None


def parse_dataset(config: dict, excel_path: Path, cache: dict) -> dict:
    materials = config.get("materials", [])
    pandas_needed = {m["sheet"] for m in materials if m.get("parserType", "old") in ("old", "new")}
    need_openpyxl = any(m.get("parserType") in ("formula", "value") for m in materials)

    sheets = {}
    wb_formulas = None
    wb_values = None
    payload = {"data": []}

    try:
        if pandas_needed:
            logger.info(f"Загружаем листы pandas: {pandas_needed}")
       
            sheet_list = list(pandas_needed)
            raw_sheets = pd.read_excel(str(excel_path), sheet_name=sheet_list, header=[0, 1])
            
            if isinstance(raw_sheets, dict):
                sheets = raw_sheets
                logger.info(f"Загружены листы (dict): {list(sheets.keys())}")
            else:
                sheets = {sheet_list[0]: raw_sheets} if sheet_list else {}
                logger.info(f"Загружен один лист: {sheet_list}")
                
            for sheet_name, df in sheets.items():
                if not isinstance(df, pd.DataFrame):
                    logger.warning(f"{sheet_name} не DataFrame, это {type(df)}")

        if need_openpyxl:
            logger.info(f"Загружаем openpyxl workbooks (formula/value материалы нужны)...")
            wb_formulas = openpyxl.load_workbook(excel_path, data_only=False)
            wb_values = openpyxl.load_workbook(excel_path, data_only=True)
            logger.info(f"openpyxl загружены. Листы: {wb_values.sheetnames}")

        for material in materials:
            ptype = material.get("parserType", "old")
            sheet_name = material.get("sheet")
            result = None

            try:
                if ptype == "old":
                    if sheet_name not in sheets:
                        logger.warning(f"Лист '{sheet_name}' не загружен (old). Пропуск.")
                        continue
                    result = parse_old(material, sheets)

                elif ptype == "new":
                    if sheet_name not in sheets:
                        logger.warning(f"Лист '{sheet_name}' не загружен (new). Пропуск.")
                        continue
                    result = parse_new(material, sheets)

                elif ptype in ("formula", "value"):
                    if not wb_values or sheet_name not in wb_values.sheetnames:
                        logger.warning(f"Лист '{sheet_name}' не найден (formula/value). Пропуск.")
                        continue
                    result = parse_excel_formulas(material, excel_path, wb_formulas, wb_values)

                else:
                    raise ValueError(f"Неизвестный parserType: {ptype}")

                if result:
                    m_id = str(result["materialId"])
                    m_date = result["dateValues"][0]["date"]

                    if cache.get(m_id) == m_date:  # debounce
                        logger.info(f"[SKIP] Material {m_id} уже отправлен на {m_date}")
                        continue

                    cache[m_id] = m_date
                    payload["data"].append(result)
                    logger.info(f"Material {m_id} обработан (date={m_date})")
                else:
                    pass

            except Exception as e:
                logger.error(f"Ошибка обработки material {material.get('materialId')}: {e}")
                logger.error(traceback.format_exc())
                continue

        return payload

    except Exception as e:
        logger.error(f"Критическая ошибка в parse_dataset: {e}")
        logger.error(traceback.format_exc())
        raise

    finally:
        if wb_formulas is not None:
            try:
                wb_formulas.close()
            except Exception as e:
                logger.warning(f"Ошибка при закрытии wb_formulas: {e}")

        if wb_values is not None:
            try:
                wb_values.close()
            except Exception as e:
                logger.warning(f"Ошибка при закрытии wb_values: {e}")



def send_payload(payload: dict, config: dict):
    server = os.getenv("SERVER_URL") or config["server"]["url"]
    route = os.getenv("SERVER_ROUTE") or config["server"]["route"]
    url = server.rstrip("/") + "/" + route.lstrip("/")
    logger.info(f"Отправка на {url}")
    response = requests.post(url, json=payload, timeout=10)
    logger.info(f"Ответ сервера: {response.status_code} {response.text}")

if __name__ == "__main__":
    CONFIG_PATH = ROOT / "config.json"
    try:
        with CONFIG_PATH.open("r", encoding="utf-8") as f:
            config = json.load(f)
        logger.info(f"Конфиг загружен: {len(config.get('materials', []))} материалов")
    except Exception as e:
        logger.error(f"Ошибка загрузки конфига {CONFIG_PATH}: {e}")
        logger.error(traceback.format_exc())
        exit(1)

    try:
        cache = load_cache()
        logger.info(f"Кэш загружен: {len(cache)} записей")

        payload = parse_dataset(config, excel_path, cache)
        logger.info(f"parse_dataset завершена: {len(payload['data'])} новых записей")

        if payload["data"]:
            send_payload(payload, config)
            save_cache(cache)
            logger.info(f"Данные отправлены и кэш сохранён")
        else:
            logger.info("[INFO] Нет новых данных для отправки")
    except Exception as e:
        logger.error(f"Критическая ошибка в main: {e}")
        logger.error(traceback.format_exc())
        exit(1)
